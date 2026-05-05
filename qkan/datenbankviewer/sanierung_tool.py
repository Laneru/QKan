"""
sanierung_tool.py - Sanierungstool für QGIS-Plugin Datenbankviewer
Modularisierte Sanierungqkan-Klasse mit vollständiger Kostenberechnung und Logik.
"""

# =========================================================
# Importe
# =========================================================

import os
import json
import sqlite3
import decimal

from .db_connection import load_qkan_connection

from PyQt5 import uic, QtGui, QtCore, QtWidgets
from PyQt5.QtWidgets import (
    QDialog,
    QTableWidgetItem,
    QMessageBox,
    QTabWidget,
    QTableWidget,
    QInputDialog,
    QDialogButtonBox,
    QListWidget,
    QAbstractItemView,
    QLabel,
    QVBoxLayout,
    QLineEdit,
    QFileDialog,
)
from PyQt5.QtCore import Qt, QMimeData
from PyQt5.QtGui import QDrag, QBrush, QColor

from qgis.utils import iface
from qgis.core import (
    QgsProject,
    QgsExpression,
    QgsExpressionContext,
    QgsExpressionContextUtils,
    QgsFeatureRequest,
)


# =========================================================
# Externe Tools
# =========================================================

try:
    from .kostenermittlung_erneuerung import Kostenermittlung_Erneuerung
except ImportError:
    KostenermittlungErneuerung = None


FORMCLASS_Sanierungstool, _ = uic.loadUiType(
    os.path.join(os.path.dirname(__file__), "Sanierungstool.ui")
)


# =========================================================
# Hauptklasse Sanierungstool
# =========================================================

class Sanierungstool(QDialog, FORMCLASS_Sanierungstool):
    """Vollständige Klasse für Sanierungsplanung und Kostenermittlung."""

    # -----------------------------------------------------
    # Konstruktor
    # -----------------------------------------------------
    def __init__(
        self,
        data,
        data_schacht,
        headers_haltungen,
        headers_schacht,
        untersuchhal,
        untersuchtag,
        untersuchtag_schacht,
        headers_gal=None,
        data_gal=None,
        planungsname=None,
        planungs_typ=None,
        parent=None,
        add_additional_columns=True,
        kosten_haltung_sanierung=None,
        kosten_haltung_renovierung=None,
        kosten_schacht_sanierung=None,
        spatialite_conn=None,
    ):
        super(Sanierungstool, self).__init__(parent)
        self.setupUi(self)

        self.setWindowModality(Qt.NonModal)
        self.setWindowFlags(
            self.windowFlags()
            | Qt.Window
            | Qt.WindowMinimizeButtonHint
            | Qt.WindowMaximizeButtonHint
        )

        self.conn = None
        self.cur = None
        self.is_spatialite = True
        self.spatialite_conn = spatialite_conn

        self._init_db_connection()

        self.untersuchtag = untersuchtag
        self.untersuchhal = untersuchhal
        self.untersuchtag_schacht = untersuchtag_schacht

        self.planungsname = planungsname
        self.planungs_typ = planungs_typ

        self.kosten_haltung_sanierung = kosten_haltung_sanierung
        self.kosten_haltung_renovierung = kosten_haltung_renovierung
        self.kosten_schacht_sanierung = kosten_schacht_sanierung

        # ---------------------------------------------
        # Signalverbindungen
        # ---------------------------------------------
        self.Berechnung_Renovierung.clicked.connect(self.Berechnung_Renovierungskosten)
        self.Schadenstabelle.itemChanged.connect(self.Berechnung)
        self.Schadenstabelle_schacht.itemChanged.connect(self.Berechnung_schacht)
        self.save_Sanierung.clicked.connect(self.speichere_sanierung)
        self.Erneuerung.clicked.connect(self.open_erneuerung)
        self.Export_Renovierung.clicked.connect(self.excel_export)
        self.Export_Renovierung_schacht.clicked.connect(self.excel_export)
        self.Export_Renovierung_gal.clicked.connect(self.excel_export)

        # =====================================================
        # 1. Verarbeitung der Haltungsdaten (Schadenstabelle)
        # =====================================================
        self.Schadenstabelle.setRowCount(len(data))

        if add_additional_columns:
            num_additional_columns = 4
            headers_with_additional_haltungen = headers_haltungen + [
                "Sanierungsverfahren 1",
                "Sanierungsverfahren 2",
                "Sanierungsverfahren 3",
                "Anzahl",
            ]
        else:
            headers_with_additional_haltungen = [
                "station",
                "kuerzel",
                "langtext",
                "charakt1",
                "charakt2",
                "quantnr1",
                "quantnr2",
                "pos_von",
                "pos_bis",
                "zd",
                "zs",
                "zb",
                "Sanierungsverfahren 1",
                "Sanierungsverfahren 2",
                "Sanierungsverfahren 3",
                "Anzahl",
            ]

        # Spaltenanzahl anhand der Header setzen
        self.Schadenstabelle.setColumnCount(len(headers_with_additional_haltungen))
        self.Schadenstabelle.setHorizontalHeaderLabels(headers_with_additional_haltungen)

        # Daten einfügen
        for i, row in enumerate(data):
            # Zeile auffüllen, falls row weniger Spalten hat (für Zusatzspalten)
            padded_row = list(row) + [""] * (
                len(headers_with_additional_haltungen) - len(row)
            )

            for j, value in enumerate(padded_row):
                item = QTableWidgetItem(str(value) if value is not None else "")
                self.Schadenstabelle.setItem(i, j, item)

                # Farblogik für ZD, ZS, ZB
                if j < len(headers_with_additional_haltungen):
                    column_name = headers_with_additional_haltungen[j]
                    if column_name.lower() in ("zd", "zs", "zb"):
                        try:
                            val_str = str(value).strip()
                            int_value = (
                                int(float(val_str))
                                if val_str and val_str.replace(".", "", 1).isdigit()
                                else None
                            )

                            brush = QtGui.QBrush(QtGui.QColor("darkgreen"))  # Default
                            if int_value == 0:
                                brush = QtGui.QBrush(QtGui.QColor("red"))
                            elif int_value == 1:
                                brush = QtGui.QBrush(QtGui.QColor("yellow"))
                            elif int_value == 2:
                                brush = QtGui.QBrush(QtGui.QColor("blue"))
                            elif int_value == 3:
                                brush = QtGui.QBrush(QtGui.QColor("lightgreen"))
                            elif int_value == 4:
                                brush = QtGui.QBrush(QtGui.QColor("green"))

                            item.setForeground(brush)
                        except (ValueError, TypeError):
                            item.setForeground(QtGui.QBrush(QtGui.QColor("black")))

        self.Schadenstabelle.viewport().update()

        # =====================================================
        # 2. Verarbeitung der Schachtdaten (Schadenstabelle_schacht)
        # =====================================================
        if data_schacht:
            self.Schadenstabelle_schacht.setRowCount(len(data_schacht))

            if add_additional_columns:
                headers_with_additional_schacht = headers_schacht + [
                    "Sanierungsverfahren 1",
                    "Sanierungsverfahren 2",
                    "Sanierungsverfahren 3",
                    "Anzahl",
                ]
            else:
                headers_with_additional_schacht = [
                    "station",
                    "kuerzel",
                    "langtext",
                    "charakt1",
                    "charakt2",
                    "quantnr1",
                    "quantnr2",
                    "pos_von",
                    "pos_bis",
                    "zd",
                    "zs",
                    "zb",
                    "Sanierungsverfahren 1",
                    "Sanierungsverfahren 2",
                    "Sanierungsverfahren 3",
                    "Anzahl",
                ]

            self.Schadenstabelle_schacht.setColumnCount(
                len(headers_with_additional_schacht)
            )
            self.Schadenstabelle_schacht.setHorizontalHeaderLabels(
                headers_with_additional_schacht
            )

            for i, row in enumerate(data_schacht):
                padded_row = list(row) + [""] * (
                    len(headers_with_additional_schacht) - len(row)
                )

                for j, value in enumerate(padded_row):
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    self.Schadenstabelle_schacht.setItem(i, j, item)

                    if j < len(headers_with_additional_schacht):
                        column_name = headers_with_additional_schacht[j]
                        if column_name.lower() in ("zd", "zs", "zb"):
                            try:
                                val_str = str(value).strip()
                                int_value = (
                                    int(float(val_str))
                                    if val_str and val_str.replace(".", "", 1).isdigit()
                                    else None
                                )

                                brush = QtGui.QBrush(QtGui.QColor("darkgreen"))
                                if int_value == 0:
                                    brush = QtGui.QBrush(QtGui.QColor("red"))
                                elif int_value == 1:
                                    brush = QtGui.QBrush(QtGui.QColor("yellow"))
                                elif int_value == 2:
                                    brush = QtGui.QBrush(QtGui.QColor("blue"))
                                elif int_value == 3:
                                    brush = QtGui.QBrush(QtGui.QColor("lightgreen"))
                                elif int_value == 4:
                                    brush = QtGui.QBrush(QtGui.QColor("green"))

                                item.setForeground(brush)
                            except (ValueError, TypeError):
                                item.setForeground(QtGui.QBrush(QtGui.QColor("black")))

        # =====================================================
        # 3. Verarbeitung der GAL-Daten (Schadenstabelle_gal)
        # =====================================================
        if data_gal is not None:
            self.Schadenstabelle_gal.setRowCount(len(data_gal))

            if add_additional_columns:
                headers_with_additional_gal = headers_gal + [
                    "Sanierungsverfahren 1",
                    "Sanierungsverfahren 2",
                    "Sanierungsverfahren 3",
                    "Anzahl",
                ]
            else:
                headers_with_additional_gal = [
                    "station",
                    "kuerzel",
                    "langtext",
                    "charakt1",
                    "charakt2",
                    "quantnr1",
                    "quantnr2",
                    "pos_von",
                    "pos_bis",
                    "zd",
                    "zs",
                    "zb",
                    "Sanierungsverfahren 1",
                    "Sanierungsverfahren 2",
                    "Sanierungsverfahren 3",
                    "Anzahl",
                ]

            self.Schadenstabelle_gal.setColumnCount(len(headers_with_additional_gal))
            self.Schadenstabelle_gal.setHorizontalHeaderLabels(
                headers_with_additional_gal
            )

            for i, row in enumerate(data_gal):
                padded_row = list(row) + [""] * (
                    len(headers_with_additional_gal) - len(row)
                )

                for j, value in enumerate(padded_row):
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    self.Schadenstabelle_gal.setItem(i, j, item)

                    if j < len(headers_with_additional_gal):
                        column_name = headers_with_additional_gal[j]
                        if column_name.lower() in ("zd", "zs", "zb"):
                            try:
                                val_str = str(value).strip()
                                int_value = (
                                    int(float(val_str))
                                    if val_str and val_str.replace(".", "", 1).isdigit()
                                    else None
                                )

                                brush = QtGui.QBrush(QtGui.QColor("darkgreen"))
                                if int_value == 0:
                                    brush = QtGui.QBrush(QtGui.QColor("red"))
                                elif int_value == 1:
                                    brush = QtGui.QBrush(QtGui.QColor("yellow"))
                                elif int_value == 2:
                                    brush = QtGui.QBrush(QtGui.QColor("blue"))
                                elif int_value == 3:
                                    brush = QtGui.QBrush(QtGui.QColor("lightgreen"))
                                elif int_value == 4:
                                    brush = QtGui.QBrush(QtGui.QColor("green"))

                                item.setForeground(brush)
                            except (ValueError, TypeError):
                                item.setForeground(QtGui.QBrush(QtGui.QColor("black")))

        # =====================================================
        # 4. Initialisierungen (Abfragen, Drag&Drop, Feature-Daten)
        # =====================================================

        # Preislisten laden
        self.Abfrage_Sanierungsverfahren()
        self.Abfrage_Sanierungsverfahren_schacht()
        # self.AbfrageSanierungsverfahrengal()  # Falls separat vorhanden

        # Drag & Drop Einstellungen
        self.Sanierungsverfahren.setDragEnabled(True)
        self.Sanierungsverfahren.setSelectionMode(QTableWidget.SingleSelection)
        self.Schadenstabelle.setAcceptDrops(True)
        self.Schadenstabelle.setSelectionMode(QTableWidget.SingleSelection)

        self.Sanierungsverfahren_schacht.setDragEnabled(True)
        self.Sanierungsverfahren_schacht.setSelectionMode(QTableWidget.SingleSelection)
        self.Schadenstabelle_schacht.setAcceptDrops(True)
        self.Schadenstabelle_schacht.setSelectionMode(QTableWidget.SingleSelection)

        self.Sanierungsverfahren_gal.setDragEnabled(True)
        self.Sanierungsverfahren_gal.setSelectionMode(QTableWidget.SingleSelection)
        self.Schadenstabelle_gal.setAcceptDrops(True)
        self.Schadenstabelle_gal.setSelectionMode(QTableWidget.SingleSelection)

        # Initialwerte für Dimension/Laenge aus Datenbankviewer übernehmen
        from .__init__ import databaseviewer
        test_dialog = databaseviewer()
        self.Dimension_Renovierung.setText(test_dialog.Dimension.text())
        self.Laenge_Renovierung.setText(test_dialog.Laenge.text())

        # ---------------------------------------------
        # Initiale Haltungs- / GAL-Infos
        # ---------------------------------------------
        if self.planungsname is None:
            # Aufruf aus Datenbankviewer
            layer = iface.activeLayer()
            if layer is None:
                self.Haltungsname.setText("No active layer.")
                self.film_dateiname.setText("")
                return

            selected_ids = layer.selectedFeatureIds()
            if len(selected_ids) == 0:
                self.Haltungsname.setText("No feature selected.")
                self.film_dateiname.setText("")
                return

            selected_feature_id = selected_ids[0]
            feature = layer.getFeature(selected_feature_id)
            self.fetch_anzahl_gal_from_layer(selected_feature_id)
        else:
            # Aufruf aus Sanierungsplaner
            if hasattr(self, "Haltungsname"):
                self.Haltungsname.setText(str(self.untersuchhal))

            layer = iface.activeLayer()
            if layer is not None:
                selected_ids = layer.selectedFeatureIds()
                if selected_ids:
                    selected_feature_id = selected_ids[0]
                    try:
                        self.fetch_anzahl_gal_from_layer(selected_feature_id)
                    except Exception as e:
                        print("Fehler bei fetch_anzahl_gal_from_layer im Planer-Aufruf:", e)

        # Gespeicherte Kosten in UI übernehmen
        try:
            if self.kosten_haltung_renovierung is not None:
                self.Renovierungskosten_2.setText(
                    f"{float(self.kosten_haltung_renovierung):.2f} €"
                )
            if self.kosten_haltung_sanierung is not None:
                self.Sanierungskosten_2.setText(
                    f"{float(self.kosten_haltung_sanierung):.2f} €"
                )
            if self.kosten_schacht_sanierung is not None:
                self.Sanierungskosten_schacht_2.setText(
                    f"{float(self.kosten_schacht_sanierung):.2f} €"
                )
        except Exception as e:
            print("Fehler beim Setzen der Kostenfelder aus Planung:", e)

        # =====================================================
        # 5. Datenbankverbindung / DB-Typ-Wahl
        # =====================================================

    # -----------------------------------------------------
    # DB-Verbindung initialisieren
    # -----------------------------------------------------
    def _init_db_connection(self):
        """
        Nutzt ausschließlich die aktive QKan-/SpatiaLite-Verbindung.
        Falls bereits eine Verbindung übergeben wurde, wird diese verwendet.
        Andernfalls wird sie über db_connection.py geladen.
        """
        self.conn = None
        self.cur = None
        self.is_spatialite = True

        print("DEBUG DB: _init_db_connection gestartet (QKan/SpatiaLite only)")

        try:
            if self.spatialite_conn is not None:
                self.conn = self.spatialite_conn
                print("DEBUG DB: vorhandene SpatiaLite-Verbindung übernommen")
            else:
                self.conn = load_qkan_connection(self)
                print("DEBUG DB: Verbindung über load_qkan_connection(self) geladen")

            if self.conn is None:
                raise RuntimeError("Keine aktive QKan-SpatiaLite-Verbindung verfügbar.")

            self.cur = self.conn.cursor()
            self.is_spatialite = True

            print("DEBUG DB: QKan-/SpatiaLite-Verbindung erfolgreich initialisiert")

        except Exception as e:
            print(f"DEBUG DB: Fehler in _init_db_connection: {e}")
            QMessageBox.warning(
                self,
                "Datenbankfehler",
                f"Fehler beim Aufbau der QKan-/SpatiaLite-Verbindung:\n{e}",
            )
            self.conn = None
            self.cur = None
            self.is_spatialite = True

    # =====================================================
    # 6. Ermittlung Anzahl offener Anschlüsse
    # =====================================================

    def fetch_anzahl_gal_from_db(self, haltnam):
        if not getattr(self, "cur", None):
            print("DEBUG GAL: Kein Cursor (self.cur ist None)")
            return 0

        print(f"DEBUG GAL: starte Abfrage für haltnam = {haltnam!r}")
        print(f"DEBUG GAL: is_spatialite = {getattr(self, 'is_spatialite', None)!r}")

        try:
            if self.is_spatialite:
                query = """
                    SELECT COUNT(*) as gal_anzahl
                    FROM untersuchdat_haltung u
                    JOIN (
                        SELECT untersuchhal, MAX(untersuchtag) AS max_tag
                        FROM untersuchdat_haltung
                        WHERE untersuchhal = ?
                        GROUP BY untersuchhal
                    ) latest
                      ON u.untersuchhal = latest.untersuchhal
                     AND u.untersuchtag = latest.max_tag
                    WHERE u.kuerzel = ?
                      AND u.charakt2 = ?
                """
                params = (haltnam, "BCA", "A")
            else:
                query = """
                    SELECT COUNT(*) as gal_anzahl
                    FROM untersuchdat_haltung u
                    JOIN (
                        SELECT untersuchhal, MAX(untersuchtag) AS max_tag
                        FROM untersuchdat_haltung
                        WHERE untersuchhal = %s
                        GROUP BY untersuchhal
                    ) latest
                      ON u.untersuchhal = latest.untersuchhal
                     AND u.untersuchtag = latest.max_tag
                    WHERE u.kuerzel = %s
                      AND u.charakt2 = %s
                """
                params = (haltnam, "BCA", "A")

            print("DEBUG GAL: Query:")
            print(query)
            print(f"DEBUG GAL: Params: {params!r}")

            self.cur.execute(query, params)
            row = self.cur.fetchone()

            print(f"DEBUG GAL: fetchone() -> {row!r}")

            count = row[0] if row else 0
            print(f"DEBUG GAL: Rückgabewert gal_anzahl = {count}")
            return count

        except Exception as e:
            print(f"DEBUG GAL: Exception in fetch_anzahl_gal_from_db: {e}")
            if getattr(self, "conn", None):
                self.conn.rollback()
            QMessageBox.warning(
                self,
                "Datenbankfehler",
                f"Fehler bei GAL-Abfrage: {e}",
            )
            return 0

    def fetch_anzahl_gal_from_layer(self, selected_feature_id):
        try:
            layer = iface.activeLayer()
            if not layer:
                raise Exception("Kein aktiver Layer gefunden.")

            print(f"DEBUG GAL: Aktiver Layer = {layer.name()}")

            selected_feature = layer.getFeature(selected_feature_id)
            if not selected_feature:
                raise Exception("Das ausgewählte Objekt wurde nicht gefunden.")

            print(f"DEBUG GAL: selected_feature.id() = {selected_feature.id()}")
            print(
                "DEBUG GAL: Felder im aktiven Layer:",
                [f.name() for f in layer.fields()],
            )

            if "haltnam" not in layer.fields().names():
                raise Exception("Feld 'haltnam' fehlt im aktiven Layer.")

            haltnam_value = selected_feature["haltnam"]
            print(f"DEBUG GAL: Haltnam Wert aus Feature = {haltnam_value!r}")

            gal_count = self.fetch_anzahl_gal_from_db(haltnam_value)
            print(f"DEBUG GAL: Ergebnis aus DB = {gal_count}")

            self.Anzahl_Stutzen.setText(str(gal_count))

        except Exception as e:
            print(f"DEBUG GAL: Exception in fetch_anzahl_gal_from_layer: {e}")
            QtWidgets.QMessageBox.critical(
                self,
                "Fehler",
                f"Fehler beim Abrufen der GAL-Anzahl: {e}",
            )

    # =====================================================
    # 7. Preislisten / Kostenabfragen
    # =====================================================

    def Abfrage_Sanierungsverfahren(self):
        preisliste_file_path = self._get_preisliste_path()

        try:
            with open(preisliste_file_path, "r", encoding="utf-8") as file:
                data = json.load(file)
        except FileNotFoundError:
            self.show_error_message(
                f"Die Datei '{preisliste_file_path}' wurde nicht gefunden. "
                f"Bitte legen Sie 'preisliste_sanierung.json' im Ordner des Tools ab."
            )
            return
        except json.JSONDecodeError:
            self.show_error_message(
                f"Die Datei '{preisliste_file_path}' enthält ungültiges JSON. "
                f"Bitte die Datei überprüfen."
            )
            return

        try:
            self.Sanierungsverfahren.setColumnCount(3)
            self.Sanierungsverfahren.setHorizontalHeaderLabels(
                ["Kategorie", "Position", "Preis"]
            )

            data_sorted = sorted(data, key=lambda x: x["Kategorie"])

            self.Sanierungsverfahren.setRowCount(len(data_sorted))
            for row_idx, item in enumerate(data_sorted):
                category_item = QTableWidgetItem(item["Kategorie"])
                position_item = QTableWidgetItem(item["Position"])
                price_item = QTableWidgetItem("{:.2f} €".format(item["Preis"]))

                self.Sanierungsverfahren.setItem(row_idx, 0, category_item)
                self.Sanierungsverfahren.setItem(row_idx, 1, position_item)
                self.Sanierungsverfahren.setItem(row_idx, 2, price_item)

            self.Sanierungsverfahren.resizeColumnsToContents()

        except Exception as e:
            self.show_error_message(f"Fehler beim Verarbeiten der JSON-Daten: {e}")


    def Abfrage_Sanierungsverfahren_schacht(self):
        print("Abfrage erfolgt")
        preisliste_file_path = self._get_preisliste_path()

        try:
            with open(preisliste_file_path, "r", encoding="utf-8") as file:
                data = json.load(file)
        except FileNotFoundError:
            self.show_error_message(
                f"Die Datei '{preisliste_file_path}' wurde nicht gefunden. "
                f"Bitte legen Sie 'preisliste_sanierung.json' im Ordner des Tools ab."
            )
            return
        except json.JSONDecodeError:
            self.show_error_message(
                f"Die Datei '{preisliste_file_path}' enthält ungültiges JSON. "
                f"Bitte die Datei überprüfen."
            )
            return

        try:
            self.Sanierungsverfahren_schacht.setColumnCount(3)
            self.Sanierungsverfahren_schacht.setHorizontalHeaderLabels(
                ["Kategorie", "Position", "Preis"]
            )

            data_sorted = sorted(data, key=lambda x: x["Kategorie"])

            self.Sanierungsverfahren_schacht.setRowCount(len(data_sorted))
            for row_idx, item in enumerate(data_sorted):
                category_item = QTableWidgetItem(item["Kategorie"])
                position_item = QTableWidgetItem(item["Position"])
                price_item = QTableWidgetItem("{:.2f} €".format(item["Preis"]))

                self.Sanierungsverfahren_schacht.setItem(row_idx, 0, category_item)
                self.Sanierungsverfahren_schacht.setItem(row_idx, 1, position_item)
                self.Sanierungsverfahren_schacht.setItem(row_idx, 2, price_item)

            self.Sanierungsverfahren_schacht.resizeColumnsToContents()

        except Exception as e:
            self.show_error_message(f"Fehler beim Verarbeiten der JSON-Daten: {e}")

    # =====================================================
    # 8. Berechnungen Haltungen / Schächte / Renovierung
    # =====================================================

    def Berechnung(self):
        sum_of_costs = 0
        sanierung_kosten = {}

        for i in range(self.Schadenstabelle.rowCount()):
            for j in range(self.Schadenstabelle.columnCount()):
                header_item = self.Schadenstabelle.horizontalHeaderItem(j)
                header_label = header_item.text() if header_item is not None else ""
                if header_label.startswith("Sanierungsverfahren"):
                    item = self.Schadenstabelle.item(i, j)
                    if item is not None:
                        sanierungsverfahren = item.text()
                        if sanierungsverfahren not in sanierung_kosten:
                            kosten = self.Abfrage_Kosten(sanierungsverfahren)
                            sanierung_kosten[sanierungsverfahren] = kosten
                        kosten = sanierung_kosten[sanierungsverfahren]

                        if header_label == "Sanierungsverfahren 3":
                            anzahl_item = self.Schadenstabelle.item(
                                i, self.Schadenstabelle.columnCount() - 1
                            )
                            anzahl_text = (
                                anzahl_item.text() if anzahl_item is not None else "1"
                            )
                            if anzahl_text:
                                anzahl = float(anzahl_text)
                                kosten *= anzahl

                        sum_of_costs += kosten

        self.Sanierungskosten.setText("{:.2f} €".format(sum_of_costs))
        sanierungskosten_2 = decimal.Decimal(sum_of_costs) * decimal.Decimal("1.19")
        self.Sanierungskosten_2.setText("{:.2f} €".format(sanierungskosten_2))
        return sanierungskosten_2

    def Berechnung_schacht(self):
        sum_of_costs = 0
        sanierung_kosten = {}

        for i in range(self.Schadenstabelle_schacht.rowCount()):
            for j in range(self.Schadenstabelle_schacht.columnCount()):
                header_item = self.Schadenstabelle_schacht.horizontalHeaderItem(j)
                header_label = header_item.text() if header_item is not None else ""
                if header_label.startswith("Sanierungsverfahren"):
                    item = self.Schadenstabelle_schacht.item(i, j)
                    if item is not None:
                        sanierungsverfahren = item.text()
                        if sanierungsverfahren not in sanierung_kosten:
                            kosten = self.Abfrage_Kosten_schacht(sanierungsverfahren)
                            sanierung_kosten[sanierungsverfahren] = kosten
                        kosten = sanierung_kosten[sanierungsverfahren]

                        if header_label == "Sanierungsverfahren 3":
                            anzahl_item = self.Schadenstabelle_schacht.item(
                                i,
                                self.Schadenstabelle_schacht.columnCount() - 1,
                            )
                            anzahl_text = (
                                anzahl_item.text() if anzahl_item is not None else "1"
                            )
                            if anzahl_text:
                                anzahl = float(anzahl_text)
                                kosten *= anzahl

                        sum_of_costs += kosten

        self.Sanierungskosten_schacht.setText("{:.2f} €".format(sum_of_costs))
        sanierungskosten_2 = decimal.Decimal(sum_of_costs) * decimal.Decimal("1.19")
        self.Sanierungskosten_schacht_2.setText(
            "{:.2f} €".format(sanierungskosten_2)
        )
        return sanierungskosten_2

    def Abfrage_Kosten(self, sanierungsverfahren):
        kosten = 0
        try:
            for i in range(self.Sanierungsverfahren.rowCount()):
                position_item = self.Sanierungsverfahren.item(i, 1)
                price_item = self.Sanierungsverfahren.item(i, 2)
                if position_item and position_item.text() == sanierungsverfahren:
                    kosten = float(price_item.text().replace(" €", ""))
                    break
        except Exception as e:
            print(
                "Fehler beim Abrufen der Kosten für das Sanierungsverfahren:",
                sanierungsverfahren,
            )
            print(e)
        return kosten

    def Abfrage_Kosten_schacht(self, sanierungsverfahren):
        kosten = 0
        try:
            for i in range(self.Sanierungsverfahren_schacht.rowCount()):
                position_item = self.Sanierungsverfahren_schacht.item(i, 1)
                price_item = self.Sanierungsverfahren_schacht.item(i, 2)
                if position_item and position_item.text() == sanierungsverfahren:
                    kosten = float(price_item.text().replace(" €", ""))
                    break
        except Exception as e:
            print(
                "Fehler beim Abrufen der Kosten für das Sanierungsverfahren:",
                sanierungsverfahren,
            )
            print(e)
        return kosten

    def Berechnung_Renovierungskosten(self):
        inlinerschlauch_kosten_text = self.Abfrage_Kosten(
            "Inlinerschlauch DN" + self.Dimension_Renovierung.text()
        )
        einmessen_stutzen_kosten_text = self.Abfrage_Kosten(
            "Einmessen der Stutzen"
        )
        hutprofil_kosten_text = self.Abfrage_Kosten(
            "Stutzen Hutprofil DN" + self.Dimension_Renovierung.text()
        )
        schachteinbindung_kosten_text = self.Abfrage_Kosten(
            "Schachteinbindung Harz/Moertel"
        )
        stutzen_oeffnen_kosten_text = self.Abfrage_Kosten(
            "Stutzen oeffnen DN" + self.Dimension_Renovierung.text()
        )
        wasserhaltung_kosten_text = self.Abfrage_Kosten("Wasserhaltung")

        # Anzahl Stutzen prüfen
        anzahl_stutzen_text = self.Anzahl_Stutzen.text()
        anzahl_stutzen = (
            int(anzahl_stutzen_text)
            if anzahl_stutzen_text and anzahl_stutzen_text.isdigit()
            else 0
        )

        laenge_renovierung = float(self.Laenge_Renovierung.text())

        inlinerschlauch_kosten = (
            float(inlinerschlauch_kosten_text) if inlinerschlauch_kosten_text else 0.0
        )
        einmessen_stutzen_kosten = (
            float(einmessen_stutzen_kosten_text)
            if einmessen_stutzen_kosten_text
            else 0.0
        )
        hutprofil_kosten = (
            float(hutprofil_kosten_text) if hutprofil_kosten_text else 0.0
        )
        schachteinbindung_kosten = (
            float(schachteinbindung_kosten_text)
            if schachteinbindung_kosten_text
            else 0.0
        )
        stutzen_oeffnen_kosten = (
            float(stutzen_oeffnen_kosten_text)
            if stutzen_oeffnen_kosten_text
            else 0.0
        )
        wasserhaltung_kosten = (
            float(wasserhaltung_kosten_text) if wasserhaltung_kosten_text else 0.0
        )

        renovierungskosten = (
            inlinerschlauch_kosten * laenge_renovierung
            + einmessen_stutzen_kosten * anzahl_stutzen
            + hutprofil_kosten * anzahl_stutzen
            + schachteinbindung_kosten * 2
            + stutzen_oeffnen_kosten * anzahl_stutzen
            + wasserhaltung_kosten * 24
        )
        renovierungskosten_2 = renovierungskosten * 1.19

        self.Renovierungskosten.setText("{:.2f} €".format(renovierungskosten))
        self.Renovierungskosten_2.setText("{:.2f} €".format(renovierungskosten_2))

        # BBA/BAG zählen
        bba_count = 0
        bag_count = 0
        atv_kuerzel_index = None

        for col in range(self.Schadenstabelle.columnCount()):
            header_item = self.Schadenstabelle.horizontalHeaderItem(col)
            if header_item and header_item.text() == "kuerzel":
                atv_kuerzel_index = col
                break

        if atv_kuerzel_index is not None:
            for row in range(self.Schadenstabelle.rowCount()):
                item = self.Schadenstabelle.item(row, atv_kuerzel_index)
                if item is not None:
                    atv_kuerzel_text = item.text()
                    if "BBA" in atv_kuerzel_text:
                        bba_count += 1
                    if "BAG" in atv_kuerzel_text:
                        bag_count += 1

        if bba_count > 0:
            QMessageBox.information(
                self,
                "Hinweis",
                f"Aufgrund von {bba_count} Wurzeleinwüchsen sind "
                f"voraussichtlich {bba_count} Fräseinsätze notwendig.",
            )
        if bag_count > 0:
            QMessageBox.information(
                self,
                "Hinweis",
                f"Aufgrund von {bag_count} einragenden Stutzen sind "
                f"voraussichtlich {bag_count} Fräseinsätze notwendig.",
            )

        # Automatische Verfahrenseintragung anbieten
        confirm = QMessageBox.question(
            self,
            "Sanierungsverfahren hinzufügen",
            "Möchten Sie automatisch Sanierungsverfahren für die zuvor genannten Schäden eintragen?",
            QMessageBox.Yes | QMessageBox.No,
        )

        if confirm == QMessageBox.Yes:
            atv_kuerzel_index = self.find_column_index(self.Schadenstabelle, "kuerzel")
            verfahren_index = self.find_column_index(
                self.Schadenstabelle, "Sanierungsverfahren 1"
            )

            if atv_kuerzel_index is not None and verfahren_index is not None:
                for row in range(self.Schadenstabelle.rowCount()):
                    atv_kuerzel_item = self.Schadenstabelle.item(
                        row, atv_kuerzel_index
                    )
                    if atv_kuerzel_item:
                        atv_kuerzel = atv_kuerzel_item.text()
                        dimension = self.Dimension_Renovierung.text()

                        if atv_kuerzel == "BBA":
                            procedure = None
                            try:
                                if "/" in dimension:
                                    procedure = (
                                        "Fraesroboter Ei-Profil 250/375 - 600/900"
                                    )
                                else:
                                    if dimension.isdigit():
                                        dn_value = int(dimension)
                                        if 100 <= dn_value <= 800:
                                            procedure = (
                                                "Fraesroboter Kreisprofil DN 100 - 800"
                                            )
                            except ValueError as e:
                                print(
                                    f"Fehler bei der Verarbeitung von dimension '{dimension}': {e}"
                                )

                            if procedure is not None:
                                procedure_item = self.Schadenstabelle.item(
                                    row, verfahren_index
                                )
                                if not procedure_item:
                                    procedure_item = QTableWidgetItem(procedure)
                                    self.Schadenstabelle.setItem(
                                        row, verfahren_index, procedure_item
                                    )
                                else:
                                    procedure_item.setText(procedure)

                        elif atv_kuerzel == "BAG":
                            procedure = None
                            try:
                                if "/" in dimension:
                                    procedure = (
                                        "Fraesroboter Ei-Profil 250/375 - 600/900"
                                    )
                                else:
                                    if dimension.isdigit():
                                        dn_value = int(dimension)
                                        if 100 <= dn_value <= 800:
                                            procedure = (
                                                "Fraesroboter Kreisprofil DN 100 - 800"
                                            )
                            except ValueError as e:
                                print(
                                    f"Fehler bei der Verarbeitung von dimension '{dimension}': {e}"
                                )

                            if procedure is not None:
                                procedure_item = self.Schadenstabelle.item(
                                    row, verfahren_index
                                )
                                if not procedure_item:
                                    procedure_item = QTableWidgetItem(procedure)
                                    self.Schadenstabelle.setItem(
                                        row, verfahren_index, procedure_item
                                    )
                                else:
                                    procedure_item.setText(procedure)

        return renovierungskosten_2

    # =====================================================
    # 9. Hilfsfunktionen
    # =====================================================

    def find_column_index(self, table, column_name):
        """Findet den Index einer Spalte basierend auf ihrem Namen."""
        for col in range(table.columnCount()):
            header_item = table.horizontalHeaderItem(col)
            if header_item and header_item.text() == column_name:
                return col
        return None

    # =====================================================
    # 10. Automatisches Befüllen von Verfahren
    # =====================================================
    def auto_fill_procedures(self):
        """Füllt automatisch Sanierungsverfahren basierend auf Kürzeln."""
        atvkuerzelindex = self.find_column_index(self.Schadenstabelle, "kuerzel")
        verfahrenindex = self.find_column_index(self.Schadenstabelle, "Sanierungsverfahren 1")

        if atvkuerzelindex is None or verfahrenindex is None:
            return

        dimension = self.Dimension_Renovierung.text()

        for row in range(self.Schadenstabelle.rowCount()):
            kuerzelitem = self.Schadenstabelle.item(row, atvkuerzelindex)
            if not kuerzelitem:
                continue

            procedure = None
            if kuerzelitem.text() == "BBA":
                # Logik für BBA (Wurzeln)
                if "x" in dimension:  # Ei-Profil
                    procedure = "Fräsroboter Ei-Profil"
                elif dimension.isdigit():
                    dn = int(dimension)
                    if 100 <= dn <= 800:
                        procedure = "Fräsroboter Kreisprofil DN 100 - 800"

            if procedure:
                self.Schadenstabelle.setItem(
                    row,
                    verfahrenindex,
                    QTableWidgetItem(procedure),
                )

    # =====================================================
    # 11. Sanierung speichern
    # =====================================================
    def speichere_sanierung(self):
        """
        Speichert Sanierungsvorschläge in die aktive QKan-/SpatiaLite-Datenbankverbindung.
        Erstellt die Tabelle 'sanierungsplaner', falls sie nicht existiert.
        Vorhandene Sanierungen derselben Planung / desselben Objekts / desselben Typs
        werden vor dem Speichern überschrieben.
        """

        # ---------------------------------------------
        # Sicherstellen, dass eine DB-Verbindung existiert
        # ---------------------------------------------
        if self.conn is None or self.cur is None:
            QMessageBox.warning(
                self,
                "Datenbankfehler",
                "Keine aktive Datenbankverbindung im Sanierungstool.",
            )
            return

        cursor = self.cur

        # ---------------------------------------------
        # Tabelle 'sanierungsplaner' erstellen
        # ---------------------------------------------
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS sanierungsplaner (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                planungsname TEXT,
                planungs_typ TEXT,
                sanierungsname TEXT,
                typ TEXT,
                untersuchtag TEXT,
                untersuchhal TEXT,
                station TEXT,
                kuerzel TEXT,
                langtext TEXT,
                charakt1 TEXT,
                charakt2 TEXT,
                quantnr1 INTEGER,
                quantnr2 INTEGER,
                pos_von REAL,
                pos_bis REAL,
                ZD INTEGER,
                ZS INTEGER,
                ZB INTEGER,
                sanierungsverfahren_1 TEXT,
                sanierungsverfahren_2 TEXT,
                sanierungsverfahren_3 TEXT,
                anzahl INTEGER,
                sanierungskosten REAL,
                renovierungskosten REAL
            )
            """
        )

        # ---------------------------------------------
        # Sanierungsnamen abfragen
        # ---------------------------------------------
        sanierungsname, ok = QInputDialog.getText(
            self,
            "Sanierung speichern",
            "Geben Sie einen Namen für die Sanierung ein:",
        )
        if not ok or not sanierungsname.strip():
            QMessageBox.warning(
                self,
                "Abbruch",
                "Speichern der Sanierung wurde abgebrochen.",
            )
            return

        sanierungsname = sanierungsname.strip()

        # ---------------------------------------------
        # Tabellen zur Speicherung auswählen
        # ---------------------------------------------
        dialog = QDialog(self)
        dialog.setWindowTitle("Sanierung speichern")

        layout = QVBoxLayout(dialog)
        label = QLabel("Welche Tabellen sollen gespeichert werden?")
        layout.addWidget(label)

        list_widget = QListWidget()
        list_widget.setSelectionMode(QAbstractItemView.MultiSelection)
        optionen = [
            "Schadenstabelle (Haltungen)",
            "Schadenstabelle_schacht (Schächte)",
            "Schadenstabelle_gal (GAL)",
        ]
        list_widget.addItems(optionen)
        layout.addWidget(list_widget)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(button_box)

        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)

        if dialog.exec() == QDialog.Accepted:
            checked_options = [item.text() for item in list_widget.selectedItems()]
            if not checked_options:
                QMessageBox.warning(
                    self,
                    "Abbruch",
                    "Keine Tabellen zur Speicherung ausgewählt.",
                )
                return
        else:
            QMessageBox.warning(
                self,
                "Abbruch",
                "Speichern der Sanierung wurde abgebrochen.",
            )
            return

        # ---------------------------------------------
        # Hilfsfunktionen
        # ---------------------------------------------
        def parse_int(text, default=0):
            if text is None:
                return default
            value = str(text).strip()
            if not value:
                return default
            try:
                return int(float(value.replace(",", ".")))
            except (ValueError, TypeError):
                return default

        def parse_float(text, default=None):
            if text is None:
                return default
            value = str(text).strip()
            if not value:
                return default

            value = value.replace("€", "").replace(" ", "")
            if "," in value and "." in value:
                value = value.replace(".", "").replace(",", ".")
            elif "," in value:
                value = value.replace(",", ".")

            try:
                return float(value)
            except (ValueError, TypeError):
                return default

        def get_item_text(table_widget, row, col, default=""):
            item = table_widget.item(row, col)
            if item is None:
                return default
            return item.text().strip()

        def get_lineedit_value(object_name, default=0.0):
            widget = self.findChild(QLineEdit, object_name)
            if widget is None:
                return default
            return parse_float(widget.text(), default=default)

        def hole_tabellendaten(table_name, typ):
            table_widget = self.findChild(QTableWidget, table_name)
            if table_widget is None:
                return []

            daten = []
            for row in range(table_widget.rowCount()):
                daten.append(
                    {
                        "station": get_item_text(table_widget, row, 0, ""),
                        "kuerzel": get_item_text(table_widget, row, 1, ""),
                        "langtext": get_item_text(table_widget, row, 2, ""),
                        "charakt1": get_item_text(table_widget, row, 3, ""),
                        "charakt2": get_item_text(table_widget, row, 4, ""),
                        "quantnr1": parse_int(get_item_text(table_widget, row, 5, "")),
                        "quantnr2": parse_int(get_item_text(table_widget, row, 6, "")),
                        "pos_von": parse_float(get_item_text(table_widget, row, 7, ""), None),
                        "pos_bis": parse_float(get_item_text(table_widget, row, 8, ""), None),
                        "ZD": parse_int(get_item_text(table_widget, row, 9, "")),
                        "ZS": parse_int(get_item_text(table_widget, row, 10, "")),
                        "ZB": parse_int(get_item_text(table_widget, row, 11, "")),
                        "sanierungsverfahren_1": get_item_text(table_widget, row, 12, ""),
                        "sanierungsverfahren_2": get_item_text(table_widget, row, 13, ""),
                        "sanierungsverfahren_3": get_item_text(table_widget, row, 14, ""),
                        "anzahl": parse_int(get_item_text(table_widget, row, 15, ""), 0),
                        "Typ": typ,
                    }
                )
            return daten

        daten_gesamt = []
        zu_loeschende_typen = set()

        # ---------------------------------------------
        # Haltungen sammeln
        # ---------------------------------------------
        if "Schadenstabelle (Haltungen)" in checked_options:
            daten_haltungen = hole_tabellendaten("Schadenstabelle", "Haltung")
            renovierungskosten = get_lineedit_value("Renovierungskosten_2", 0.0)
            sanierungskosten = get_lineedit_value("Sanierungskosten_2", 0.0)

            for schaden in daten_haltungen:
                schaden["renovierungskosten"] = renovierungskosten
                schaden["sanierungskosten"] = sanierungskosten

            daten_gesamt.extend(daten_haltungen)
            zu_loeschende_typen.add("Haltung")

        # ---------------------------------------------
        # Schächte sammeln
        # ---------------------------------------------
        if "Schadenstabelle_schacht (Schächte)" in checked_options:
            daten_schacht = hole_tabellendaten("Schadenstabelle_schacht", "Schacht")
            sanierungskosten_schacht = get_lineedit_value("Sanierungskosten_schacht_2", 0.0)

            for schaden in daten_schacht:
                schaden["renovierungskosten"] = 0.0
                schaden["sanierungskosten"] = sanierungskosten_schacht

            daten_gesamt.extend(daten_schacht)
            zu_loeschende_typen.add("Schacht")

        # ---------------------------------------------
        # GAL sammeln
        # ---------------------------------------------
        if "Schadenstabelle_gal (GAL)" in checked_options:
            daten_gal = hole_tabellendaten("Schadenstabelle_gal", "GAL")
            renovierungskosten_gal = get_lineedit_value("Renovierungskosten_gal_2", 0.0)
            sanierungskosten_gal = get_lineedit_value("Sanierungskosten_gal_2", 0.0)

            for schaden in daten_gal:
                schaden["renovierungskosten"] = renovierungskosten_gal
                schaden["sanierungskosten"] = sanierungskosten_gal

            daten_gesamt.extend(daten_gal)
            zu_loeschende_typen.add("GAL")

        if not daten_gesamt:
            QMessageBox.warning(
                self,
                "Hinweis",
                "Es wurden keine Daten zum Speichern gefunden.",
            )
            return

        # ---------------------------------------------
        # Planungsmetadaten
        # ---------------------------------------------
        planungsname = getattr(self, "planungsname", None)
        planungs_typ = getattr(self, "planungs_typ", None)

        # ---------------------------------------------
        # Vorhandene Datensätze löschen (überschreiben)
        # ---------------------------------------------
        if planungsname and planungs_typ and zu_loeschende_typen:
            for typ_del in zu_loeschende_typen:
                cursor.execute(
                    """
                    DELETE FROM sanierungsplaner
                    WHERE planungsname = ?
                    AND planungs_typ = ?
                    AND untersuchhal = ?
                    AND typ = ?
                    """,
                    (planungsname, planungs_typ, self.untersuchhal, typ_del),
                )

        # ---------------------------------------------
        # Neue Datensätze einfügen
        # ---------------------------------------------
        insert_count = 0

        try:
            for schaden in daten_gesamt:
                if schaden["Typ"] in ("Haltung", "GAL"):
                    untersuchtag = self.untersuchtag
                elif schaden["Typ"] == "Schacht":
                    untersuchtag = self.untersuchtag_schacht
                else:
                    untersuchtag = None

                cursor.execute(
                    """
                    INSERT INTO sanierungsplaner (
                        planungsname,
                        planungs_typ,
                        sanierungsname,
                        typ,
                        untersuchtag,
                        untersuchhal,
                        station,
                        kuerzel,
                        langtext,
                        charakt1,
                        charakt2,
                        quantnr1,
                        quantnr2,
                        pos_von,
                        pos_bis,
                        ZD,
                        ZS,
                        ZB,
                        sanierungsverfahren_1,
                        sanierungsverfahren_2,
                        sanierungsverfahren_3,
                        anzahl,
                        sanierungskosten,
                        renovierungskosten
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        planungsname,
                        planungs_typ,
                        sanierungsname,
                        schaden["Typ"],
                        untersuchtag,
                        self.untersuchhal,
                        schaden["station"],
                        schaden["kuerzel"],
                        schaden["langtext"],
                        schaden["charakt1"],
                        schaden["charakt2"],
                        schaden["quantnr1"],
                        schaden["quantnr2"],
                        schaden["pos_von"],
                        schaden["pos_bis"],
                        schaden["ZD"],
                        schaden["ZS"],
                        schaden["ZB"],
                        schaden["sanierungsverfahren_1"],
                        schaden["sanierungsverfahren_2"],
                        schaden["sanierungsverfahren_3"],
                        schaden["anzahl"],
                        schaden["sanierungskosten"],
                        schaden["renovierungskosten"],
                    ),
                )
                insert_count += 1

            self.conn.commit()

        except Exception as e:
            self.conn.rollback()
            QMessageBox.critical(
                self,
                "Datenbankfehler",
                f"Fehler beim Speichern der Sanierung:\n{e}",
            )
            return

        # ---------------------------------------------
        # Rückmeldung
        # ---------------------------------------------
        if insert_count > 0:
            QMessageBox.information(
                self,
                "Erfolg",
                f"Sanierung '{sanierungsname}' wurde erfolgreich gespeichert ({insert_count} Datensätze).",
            )
        else:
            QMessageBox.warning(
                self,
                "Hinweis",
                "Es wurden keine Datensätze gespeichert.",
            )
    # =====================================================
    # 12. Objektdaten für Export
    # =====================================================
    def _get_objekt_infos_for_export(self):
        """
        Liefert (objekt_name, strasse, schacht_oben, schacht_unten,
                 dimension, laenge, material)
        immer aus dem aktuell selektierten Feature (mit Fallbacks wie im Datenbankviewer).
        """
        layer = iface.activeLayer()
        if layer is None or len(layer.selectedFeatureIds()) == 0:
            QMessageBox.warning(
                self,
                "Fehler",
                "Kein Layer oder Feature selektiert.",
            )
            return "", "", "", "", "", "", ""

        feature = layer.getFeature(layer.selectedFeatureIds()[0])

        existing_attributes = feature.fields().names()
        attribute_values = {
            attr.lower(): str(feature.attribute(attr))
            for attr in existing_attributes
            if feature.attribute(attr) is not None
        }

        def get_val(keys_wanted):
            if isinstance(keys_wanted, str):
                keys_wanted = [keys_wanted]
            for kw in keys_wanted:
                val = attribute_values.get(kw.lower())
                if val:
                    return val
            return ""

        objekt_name = get_val(["haltnam", "leitnam", "schnam", "schoben"])
        strasse = get_val("strasse")
        schacht_oben = get_val(["schoben", "schnam"])
        schacht_unten = get_val("schunten")
        material = get_val("material")

        laenge_raw = get_val("laenge")
        if laenge_raw:
            try:
                laenge = f"{float(laenge_raw.replace(',', '.')):.2f}"
            except ValueError:
                laenge = laenge_raw
        else:
            laenge = ""

        breite_str = get_val("breite")
        hoehe_str = get_val("hoehe")
        try:
            b_int = int(float(breite_str.replace(",", "."))) if breite_str else ""
            h_int = int(float(hoehe_str.replace(",", "."))) if hoehe_str else ""
            if b_int == h_int and b_int:
                dimension = str(b_int)
            elif b_int and h_int:
                dimension = f"{b_int}/{h_int}"
            else:
                dimension = get_val("dimension")
        except ValueError:
            dimension = (
                f"{breite_str}/{hoehe_str}"
                if breite_str and hoehe_str
                else get_val("dimension")
            )

        return objekt_name, strasse, schacht_oben, schacht_unten, dimension, laenge, material

    # =====================================================
    # 13. Tabellen → DataFrame-Helfer
    # =====================================================
    def _table_to_dataframe(self, table_widget, columns):
        import pandas as pd

        data_rows = []
        for row in range(table_widget.rowCount()):
            row_data = []
            for col in range(len(columns)):
                item = table_widget.item(row, col)
                row_data.append(item.text() if item is not None else "")
            data_rows.append(row_data)

        return pd.DataFrame(data_rows, columns=columns)

    # =====================================================
    # 14. Excel-Export
    # =====================================================
    def excel_export(self):
        """
        Exportiert die Schadenstabellen (Haltungen, Schächte, GAL) und Objektdaten
        in eine Excel-Vorlage (xlsm mit VBA), wobei Objektdaten immer aus dem
        aktuell selektierten QGIS-Feature kommen.
        Haltungen und Schächte werden parallel in derselben Zeile geschrieben.
        """
        import pandas as pd
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        import subprocess
        from PyQt5.QtWidgets import QFileDialog, QMessageBox

        # 1) Vorlage auswählen
        template_options = QFileDialog.Options()
        template_file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Excel-Vorlage auswählen",
            "",
            "Excel-Dateien (*.xlsm)",
            options=template_options,
        )
        if not template_file_path:
            return

        workbook = openpyxl.load_workbook(template_file_path, keep_vba=True)

        if "Tabelle1" not in workbook.sheetnames:
            QMessageBox.warning(
                self,
                "Fehler",
                "Die Vorlage enthält kein Blatt 'Tabelle1'.",
            )
            return
        if "Tabelle2" not in workbook.sheetnames:
            QMessageBox.warning(
                self,
                "Fehler",
                "Die Vorlage enthält kein Blatt 'Tabelle2'.",
            )
            return

        sheet_table1 = workbook["Tabelle1"]
        sheet_table2 = workbook["Tabelle2"]

        # 2) Schadenstabelle (Haltungen) als DataFrame
        df_haltungen = self._table_to_dataframe(
            self.Schadenstabelle,
            columns=[
                "Station",
                "Kürzel",
                "Langtext",
                "Charakt1",
                "Charakt2",
                "QuantNr1",
                "QuantNr2",
                "von",
                "bis",
                "ZD",
                "ZS",
                "ZB",
                "Sanierungsverfahren 1",
                "Sanierungsverfahren 2",
                "Sanierungsverfahren 3",
                "Anzahl",
            ],
        )

        # 3) Schacht-Export optional
        df_schacht = None
        confirm_schacht = QMessageBox.question(
            self,
            "Export Schachtuntersuchung",
            "Möchten Sie die Daten der Schachtuntersuchung ebenfalls exportieren?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if (
            confirm_schacht == QMessageBox.Yes
            and hasattr(self, "Schadenstabelle_schacht")
        ):
            df_schacht = self._table_to_dataframe(
                self.Schadenstabelle_schacht,
                columns=[
                    "Schacht_Station",
                    "Schacht_Kürzel",
                    "Schacht_Langtext",
                    "Schacht_Charakt1",
                    "Schacht_Charakt2",
                    "Schacht_QuantNr1",
                    "Schacht_QuantNr2",
                    "Schacht_von",
                    "Schacht_bis",
                    "Schacht_ZD",
                    "Schacht_ZS",
                    "Schacht_ZB",
                    "Schacht_Sanierungsverfahren 1",
                    "Schacht_Sanierungsverfahren 2",
                    "Schacht_Sanierungsverfahren 3",
                    "Schacht_Anzahl",
                ],
            )

        # 4) Haltungen und Schächte parallel schreiben
        if df_schacht is not None:
            max_rows = max(len(df_haltungen), len(df_schacht))
            if len(df_haltungen) < max_rows:
                df_haltungen = df_haltungen.reindex(range(max_rows))
            if len(df_schacht) < max_rows:
                df_schacht = df_schacht.reindex(range(max_rows))

            df_gesamt = pd.concat([df_haltungen, df_schacht], axis=1)

            for r_idx, row in enumerate(
                dataframe_to_rows(df_gesamt, index=False, header=True), start=1
            ):
                for c_idx, value in enumerate(row, start=1):
                    sheet_table1.cell(row=r_idx, column=c_idx, value=value)
        else:
            for row in dataframe_to_rows(df_haltungen, index=False, header=True):
                sheet_table1.append(row)

        # 5) GAL-Export optional
        if (
            hasattr(self, "Schadenstabelle_gal")
            and self.Schadenstabelle_gal.rowCount() > 0
        ):
            confirm_gal = QMessageBox.question(
                self,
                "Export GAL",
                "Möchten Sie die GAL-Daten ebenfalls exportieren?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if confirm_gal == QMessageBox.Yes:
                df_gal = self._table_to_dataframe(
                    self.Schadenstabelle_gal,
                    columns=[
                        "GAL_Station",
                        "GAL_Kürzel",
                        "GAL_Langtext",
                        "GAL_Charakt1",
                        "GAL_Charakt2",
                        "GAL_QuantNr1",
                        "GAL_QuantNr2",
                        "GAL_von",
                        "GAL_bis",
                        "GAL_ZD",
                        "GAL_ZS",
                        "GAL_ZB",
                        "GAL_Sanierungsverfahren 1",
                        "GAL_Sanierungsverfahren 2",
                        "GAL_Sanierungsverfahren 3",
                        "GAL_Anzahl",
                    ],
                )
                start_row = sheet_table1.max_row + 2
                for r_idx, row in enumerate(
                    dataframe_to_rows(df_gal, index=False, header=True),
                    start=start_row,
                ):
                    for c_idx, value in enumerate(row, start=1):
                        sheet_table1.cell(row=r_idx, column=c_idx, value=value)

        # 6) Tabelle2: Objektdaten und Kosten
        (
            objekt_name,
            strasse,
            schacht_oben,
            schacht_unten,
            dimension,
            laenge,
            material,
        ) = self._get_objekt_infos_for_export()

        renovierungskosten_text = (
            self.Renovierungskosten_2.text()
            if hasattr(self, "Renovierungskosten_2")
            else "0 €"
        )
        sanierungskosten_text = (
            self.Sanierungskosten_2.text()
            if hasattr(self, "Sanierungskosten_2")
            else "0 €"
        )
        sanierungskosten_schacht_text = (
            self.Sanierungskosten_schacht_2.text()
            if hasattr(self, "Sanierungskosten_schacht_2")
            else "0 €"
        )

        def to_float(text):
            if not text:
                return 0.0
            t = text.replace("€", "").replace(" ", "").strip()
            if "," in t and "." in t:
                t = t.replace(".", "").replace(",", ".")
            elif "," in t and "." not in t:
                t = t.replace(",", ".")
            try:
                return float(t)
            except ValueError:
                return 0.0

        renovierungskosten = to_float(renovierungskosten_text)
        sanierungskosten = to_float(sanierungskosten_text)
        sanierungskosten_schacht = to_float(sanierungskosten_schacht_text)

        data_row_table2 = [
            strasse,
            schacht_oben,
            schacht_unten,
            dimension,
            laenge,
            material,
            sanierungskosten,
            renovierungskosten,
            sanierungskosten_schacht,
        ]

        target_row = 1  # oder 2, wie du es brauchst

        for col_idx, value in enumerate(data_row_table2, start=1):
            sheet_table2.cell(row=target_row, column=col_idx, value=value)

        # 7) Datei speichern
        export_options = QFileDialog.Options()
        export_file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Excel-Datei speichern",
            "",
            "Excel-Dateien (*.xlsm)",
            options=export_options,
        )
        if export_file_path:
            workbook.save(export_file_path)
            QMessageBox.information(
                self,
                "Erfolg",
                f"Excel-Datei erfolgreich gespeichert:\n{export_file_path}",
            )
            try:
                subprocess.Popen(["start", "", export_file_path], shell=True)
            except Exception as e:
                print(f"Fehler beim Öffnen der Datei: {e}")

    # =====================================================
    # 15. Erneuerungsdialog öffnen
    # =====================================================
    def open_erneuerung(self):
        print("DEBUG SANIERUNG: conn =", self.conn)
        print("DEBUG SANIERUNG: cur =", self.cur)
        print("DEBUG SANIERUNG: is_spatialite =", self.is_spatialite)

        dimension_value = self.Dimension_Renovierung.text()
        laenge_value = self.Laenge_Renovierung.text()
        stutzen_value = self.Anzahl_Stutzen.text()

        dialog = Kostenermittlung_Erneuerung(
            dimension=dimension_value,
            laenge=laenge_value,
            anzahl_stutzen=stutzen_value,
            parent=self,
        )

        dialog.conn = self.conn
        dialog.cur = self.cur
        dialog.is_spatialite = self.is_spatialite

        dialog.exec_()

    # =====================================================
    # 16. Fehlerdialog
    # =====================================================
    def show_error_message(self, msg):
        QMessageBox.critical(self, "Fehler", msg)

    # =====================================================
    # 17. Drag & Drop Events
    # =====================================================
    def startDrag(self, event):
        sender = self.sender()
        item = sender.currentItem()
        if item:
            mimedata = QMimeData()
            mimedata.setText(item.text())
            drag = QDrag(self)
            drag.setMimeData(mimedata)
            drag.exec_(Qt.MoveAction)

    def dragEnterEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()

    def dragMoveEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()

    def dropEvent(self, event):
        if event.mimeData().hasText():
            # Ziel bestimmen (welche Tabelle hat den Drop empfangen?)
            # Hier vereinfacht: Standard-Logik / Platzhalter
            # Im Originalcode wird dropEvent direkt an die Tabelle gebunden
            # oder über EventFilter gelöst.
            event.acceptProposedAction()

    # =====================================================
    # 18. Hilfsfunktion: Pfad zur Preisliste
    # =====================================================

    def _get_preisliste_path(self):
        """
        Liefert den absoluten Pfad zur JSON-Preisliste im selben Ordner
        wie dieses Tool-Modul.
        """
        tool_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(tool_dir, "json","preisliste_sanierung.json")