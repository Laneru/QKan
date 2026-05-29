import os.path
import os
import sys

from qgis.utils import iface
from PyQt5 import uic, QtWidgets
from PyQt5.QtCore import QTimer, QTime, pyqtSignal, QObject, QDate, Qt
import PyQt5.QtWidgets
from PyQt5.QtSql import QSqlDatabase, QSqlQuery
from PyQt5 import QtGui
from PyQt5.QtGui import QColor
from qgis.core import QgsProject, QgsExpression, QgsExpressionContext, QgsExpressionContextUtils, QgsFeatureRequest, QgsMapLayer, QgsMessageLog
from qgis.PyQt.QtWidgets import QDialog, QFileDialog, QFrame, QPushButton, QSlider,QApplication, QTableWidgetItem, QTableWidget, QTabWidget, QVBoxLayout, QMessageBox
from qgis.PyQt.QtWidgets import (
    QCheckBox,
    QDialogButtonBox,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QPushButton,
    QRadioButton,
    QWidget,
)

from collections import defaultdict
import psycopg2
import pandas as pd
import sqlite3
import subprocess
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import locale
from datetime import datetime

FORM_CLASS_Kostenermittlung, _ = uic.loadUiType(
    os.path.join(os.path.dirname(__file__), "res", "Kostenermittlung_Reinigung_Test.ui")
)

class Kostenermittlung(QDialog, FORM_CLASS_Kostenermittlung):

    Reinigung_netto_MW = QLineEdit
    Reinigung_brutto_MW = QLineEdit
    Reinigung_netto_RW = QLineEdit
    Reinigung_brutto_RW = QLineEdit    
    Reinigung_netto_SW = QLineEdit
    Reinigung_brutto_SW = QLineEdit
    Reinigung_befahrung_netto_MW = QLineEdit
    Reinigung_befahrung_brutto_MW = QLineEdit
    Reinigung_befahrung_netto_RW = QLineEdit
    Reinigung_befahrung_brutto_RW = QLineEdit    
    Reinigung_befahrung_netto_SW = QLineEdit
    Reinigung_befahrung_brutto_SW = QLineEdit
    Befahrung_netto_MW = QLineEdit
    Befahrung_brutto_MW = QLineEdit
    Befahrung_netto_RW = QLineEdit
    Befahrung_brutto_RW = QLineEdit    
    Befahrung_netto_SW = QLineEdit
    Befahrung_brutto_SW = QLineEdit
    Auslesung = QPushButton
    Auslesung_2 = QPushButton
    Liste_Haltungen = QTableWidget
    Liste_Haltungen_2 = QTableWidget
    Berechnung_Reinigung = QPushButton
    Berechnung_TV = QPushButton
    checkBox_Panoramo = QCheckBox
    checkBox_GAL = QCheckBox
    Excel_Export = QPushButton
    importpostgresql = QPushButton
    laenge_mw = QLineEdit
    laenge_rw = QLineEdit
    laenge_sw = QLineEdit
    laenge_ges = QLineEdit
    prozent_mw = QLineEdit
    prozent_rw = QLineEdit
    prozent_sw = QLineEdit
    prozent_ges = QLineEdit
    tableWidget = QTableWidget
    tableWidget2 = QTableWidget
    Excel_Export_2 = QPushButton

    def __init__(self, parent=None):
        super(Kostenermittlung, self).__init__(parent)
        self.setupUi(self)
        self.iface = iface
        self.parent_plugin = parent # Referenz speichern

        self.reinigung_costs = {}
        self.reinigung_TV_costs = {}
        self.TV_costs = {}
        self.TV_SAT_costs = {}
        self.Panoramo_costs = {}

        # DATENBANKVERBINDUNG VOM ELTERN-PLUGIN ÜBERNEHMEN
        # Wir erstellen keine eigene Verbindung mehr, um Konflikte zu vermeiden
        # und DB-agnostisch zu bleiben.
        if hasattr(parent, 'conn') and parent.conn is not None:
            self.conn = parent.conn
            self.db_type = getattr(parent, 'db_type', 'postgres') # Default postgres
            self.is_spatialite = (self.db_type == 'spatialite')
        else:
            # Fallback: Versuche eigene Verbindung (nur Postgres, wie früher)
            # Das ist nur für den Fall, dass der Dialog standalone getestet wird
            self.is_spatialite = False 
            self.connect_db_fallback()

        self.cur = self.conn.cursor()

        self.Auslesung.clicked.connect(self.showSelectedFeatures)
        self.Auslesung_2.clicked.connect(self.showSelectedFeatures)

        self.Berechnung_Reinigung.clicked.connect(self.calculateCleaningCost)

        self.Berechnung_TV.clicked.connect(self.calculateCleaningTVCost)

        self.Excel_Export.clicked.connect(self.export_excel)
      
        # Verbinde die Checkboxen mit den Methoden
        self.checkBox_GAL.stateChanged.connect(self.on_checkbox_changed)
        self.checkBox_Panoramo.stateChanged.connect(self.on_checkbox_changed)
        
        # Setze die locale für die gewünschte Formatierung
        try:
            locale.setlocale(locale.LC_NUMERIC, 'de_DE.UTF-8')
        except:
            locale.setlocale(locale.LC_NUMERIC, '') # System default

        print(f"Kostenermittlung initiiert. DB-Typ: {'Spatialite' if self.is_spatialite else 'PostgreSQL'}")

    def connect_db_fallback(self):
        """Fallback Verbindungsmethode (nur Postgres), falls Parent keine liefert."""
        try:
            db_password_file_path = os.path.join(os.path.dirname(__file__), '..', 'settings', 'database.json')
            with open(db_password_file_path, 'r') as json_file:
                data = json.load(json_file)
                db_host = data.get('db_host')
                db_database = "Kanaldatenbank"
                db_user = data.get('db_username')
                db_password = data.get('db_password')

            self.conn = psycopg2.connect(
                host=db_host,
                database=db_database,
                user=db_user,
                password=db_password
            )
        except Exception as e:
            QMessageBox.critical(self, "Verbindungsfehler", f"Konnte DB nicht verbinden: {e}")

    def on_checkbox_changed(self):
        # Hier kannst du beide Methoden aufrufen
        self.calculateCleaningCost()
        self.calculateCleaningTVCost()

    def showSelectedFeatures(self):
        print("Test showSelectedFeatures aufgerufen")
        
        # Layer holen und selektierte Features übergeben
        active_layer = self.iface.activeLayer()
        if not active_layer:
            print("Kein aktiver Layer gefunden.")
            return

        selected_features = active_layer.selectedFeatures()
        self.clearTableWidget()
        self.populateTableWidget(selected_features)

    def clearTableWidget(self):
        self.Liste_Haltungen.clearContents()
        self.Liste_Haltungen.setRowCount(0)
        self.Liste_Haltungen.setColumnCount(0)
        self.Liste_Haltungen_2.clearContents()
        self.Liste_Haltungen_2.setRowCount(0)
        self.Liste_Haltungen_2.setColumnCount(0)

    def get_gal_counts_from_db(self, haltnams):
        gal_counts = {}
        try:
            haltnams_list = list(haltnams)
            if not haltnams_list: return {}

            # Unterscheidung Postgres / Spatialite für Array-Abfragen
            if self.is_spatialite:
                # SQLite: IN (?, ?, ...) Syntax
                placeholders = ','.join(['?'] * len(haltnams_list))
                
                # Wir müssen die Liste zweimal übergeben (einmal für Subselect, einmal für Hauptabfrage wäre komplex)
                # Einfacher: Query optimieren oder Liste duplizieren?
                # Optimierte Query für SQLite:
                query = f"""
                    SELECT u.untersuchhal, COUNT(*) as gal_anzahl
                    FROM untersuchdat_haltung u
                    JOIN (
                        SELECT untersuchhal, MAX(untersuchtag) AS max_tag
                        FROM untersuchdat_haltung
                        WHERE untersuchhal IN ({placeholders})
                        GROUP BY untersuchhal
                    ) latest ON u.untersuchhal = latest.untersuchhal AND u.untersuchtag = latest.max_tag
                    WHERE u.kuerzel = ? AND u.charakt2 = ?
                    GROUP BY u.untersuchhal
                """
                params = haltnams_list + ['BCA', 'A']
                self.cur.execute(query, params)
                
            else:
                # PostgreSQL: ANY(%s) Syntax
                query = """
                    SELECT u.untersuchhal, COUNT(*) as gal_anzahl
                    FROM untersuchdat_haltung u
                    JOIN (
                        SELECT untersuchhal, MAX(untersuchtag) AS max_tag
                        FROM untersuchdat_haltung
                        WHERE untersuchhal = ANY(%s)
                        GROUP BY untersuchhal
                    ) latest ON u.untersuchhal = latest.untersuchhal AND u.untersuchtag = latest.max_tag
                    WHERE u.kuerzel = %s AND u.charakt2 = %s
                    GROUP BY u.untersuchhal
                """
                self.cur.execute(query, (haltnams_list, 'BCA', 'A'))

            rows = self.cur.fetchall()
            gal_counts = {row[0]: row[1] for row in rows}
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.warning(self, "Datenbankfehler", f"Fehler bei GAL-Abfrage in Datenbank: {e}")
            gal_counts = {}

        return gal_counts

    def populateTableWidget(self, selected_features):
        print("Populating Table...")
        active_layer = iface.activeLayer()
        if not active_layer or active_layer.name() != "Haltungen":
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Question)
            msg_box.setWindowTitle("Layer-Auswahl")
            msg_box.setText("Der Layer 'Haltungen' ist nicht aktiv. Soll er jetzt ausgewählt werden?")
            msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            if msg_box.exec_() == QMessageBox.Yes:
                h_layer_list = QgsProject.instance().mapLayersByName("Haltungen")
                if h_layer_list:
                    h_layer = h_layer_list[0]
                    iface.setActiveLayer(h_layer)
                    selected_features = h_layer.selectedFeatures()
                    self.populateTableWidget(selected_features)
                    return
                else:
                    QMessageBox.warning(self, "Layer fehlt", "Der Layer 'Haltungen' ist im Projekt nicht vorhanden.")
                    return
            else:
                return

        selected_haltnams = {f['haltnam'] for f in selected_features}
        if not selected_haltnams:
            print("Keine Haltungsnamen ausgewählt.")
            return

        try:
            selected_haltnams_list = list(selected_haltnams)
            
            # DB-spezifische Abfrage für Haltungen
            if self.is_spatialite:
                placeholders = ','.join(['?'] * len(selected_haltnams_list))
                query = f"""
                    SELECT haltnam, laenge, entwart, hoehe, breite 
                    FROM haltungen 
                    WHERE haltnam IN ({placeholders})
                """
                self.cur.execute(query, selected_haltnams_list)
            else:
                query = """
                    SELECT haltnam, laenge, entwart, hoehe, breite 
                    FROM public.haltungen 
                    WHERE haltnam = ANY(%s)
                """
                self.cur.execute(query, (selected_haltnams_list,))

            filtered_features = self.cur.fetchall()
            print(f"Anzahl gefundener Datensätze: {len(filtered_features)}")
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.warning(self, "Datenbankfehler", f"Fehler bei SQL-Abfrage: {e}")
            return

        headers = ["Haltungsname", "Länge", "Entwässerungssystem", "Dimension", "Anzahl GALs"] + \
                ["Kosten Reinigung", "Kosten Reinigung TV", "Kosten TV", "Kosten GAL", "Kosten Panoramo"]

        self.Liste_Haltungen.setColumnCount(len(headers))
        self.Liste_Haltungen.setHorizontalHeaderLabels(headers)
        self.Liste_Haltungen.setRowCount(len(filtered_features))

        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Question)
        msg_box.setWindowTitle("Abfrage")
        msg_box.setText("Soll die Anzahl der GALs aus der Datenbank ermittelt werden?")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        fetch_gal = msg_box.exec_() == QMessageBox.Yes

        gal_counts = {}
        if fetch_gal:
            gal_counts = self.get_gal_counts_from_db(selected_haltnams)
        else:
            gal_counts = {}

        self.Liste_Haltungen.setUpdatesEnabled(False)
        self.Liste_Haltungen.blockSignals(True)

        try:
            for i, feature in enumerate(filtered_features):
                haltnam, laenge, entwart, hoehe, breite = feature
                
                # Handling falls Dimension NULL ist
                h = int(hoehe) if hoehe is not None else 0
                b = int(breite) if breite is not None else 0
                
                dimension = f"{b}/{h}" if h != b else str(h)
                anzahl_gal = gal_counts.get(haltnam, 0) if fetch_gal else "Nicht berechnet"

                values = [
                    haltnam,
                    laenge,
                    entwart,
                    dimension,
                    anzahl_gal,
                    "", "", "", "", ""
                ]

                for j, value in enumerate(values):
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    self.Liste_Haltungen.setItem(i, j, item)

        finally:
            self.Liste_Haltungen.resizeColumnsToContents()
            self.Liste_Haltungen.setUpdatesEnabled(True)
            self.Liste_Haltungen.blockSignals(False)

    def fetch_anzahl_gal_from_layer(self, haltnam_value):
        try:
            layers = QgsProject.instance().mapLayersByName("Einzelschäden_Haltungen")
            if not layers:
                print("Layer 'Einzelschäden_Haltungen' nicht gefunden.")
                return 0
            untersuchungsdaten_layer = layers[0]
            
            context = QgsExpressionContext()
            context.appendScopes(QgsExpressionContextUtils.globalProjectLayerScopes(untersuchungsdaten_layer))

            expression_latest_date = QgsExpression(f"untersuchhal = '{haltnam_value}'")
            untersuchungsdaten_layer.selectByExpression(expression_latest_date.expression())
            features = [f for f in untersuchungsdaten_layer.selectedFeatures()]
            
            if not features:
                return 0

            latest_date = max(f["untersuchtag"] for f in features)
            
            gal_count = sum(
                1 for f in features
                if f["untersuchtag"] == latest_date and f["kuerzel"] == "BCA" and f["charakt2"] == "A"
            )
            return gal_count
        except Exception as e:
            print(f"Fehler beim Abrufen der GAL-Anzahl: {e}")
            return 0

    def load_cleaning_costs(self):
        file_path = os.path.join(os.path.dirname(__file__), '..', 'settings', 'preisliste_untersuchung.json')
        if os.path.exists(file_path):
            with open(file_path, 'r', encoding='utf-8') as file:
                try:
                    data = json.load(file)
                    self.reinigung_costs = data.get('reinigung_costs', {})
                    self.reinigung_TV_costs = data.get('reinigung_TV_costs', {})
                    self.TV_costs = data.get('TV_costs', {})
                    self.TV_SAT_costs = data.get('TV_SAT_costs', {})
                    self.Panoramo_costs = data.get('Panoramo_costs', {})
                except json.JSONDecodeError:
                    print("Fehler beim Laden der JSON-Datei.")

    def calculateCleaningCost(self):
        self.load_cleaning_costs()
        dimension_column_index = 3
        length_column_index = 1
        system_column_index = 2
        cleaningcost_column_index = 5

        # Mapping für alternative Schreibweisen der Systeme
        system_mapping = {
            "Mischwasser": ["Mischwasser", "KM Mischwasserkanal"],
            "Regenwasser": ["Regenwasser", "KR Regenwasserkanal", "BA Bachverrohrung"],
            "Schmutzwasser": ["Schmutzwasser", "KS Schmutzwasserkanal"]
        }

        total_cost_mw = 0
        total_cost_rw = 0
        total_cost_sw = 0

        for row in range(self.Liste_Haltungen.rowCount()):
            dimension_item = self.Liste_Haltungen.item(row, dimension_column_index)
            length_item = self.Liste_Haltungen.item(row, length_column_index)
            system_item = self.Liste_Haltungen.item(row, system_column_index)

            if dimension_item is not None and length_item is not None and system_item is not None:
                dimension = dimension_item.text()
                length_text = length_item.text()
                system = system_item.text()

                if length_text is None or length_text.strip() == '' or length_text.lower() == 'none':
                    continue

                try:
                    length = float(length_text)
                except ValueError:
                    continue

                for key, values in system_mapping.items():
                    if system in values:
                        system = key
                        break

                if dimension in self.reinigung_costs and system in self.reinigung_costs[dimension]:
                    cost_per_unit = self.reinigung_costs[dimension][system]
                    cost = length * cost_per_unit

                    brutto_cost = cost * 1.19
                    cost_item = QTableWidgetItem("{:.2f} €".format(brutto_cost))
                    self.Liste_Haltungen.setItem(row, cleaningcost_column_index, cost_item)

                    if system == "Mischwasser":
                        total_cost_mw += cost
                    elif system == "Regenwasser":
                        total_cost_rw += cost
                    elif system == "Schmutzwasser":
                        total_cost_sw += cost

        self.Reinigung_netto_MW.setText("{:.2f} €".format(total_cost_mw))
        self.Reinigung_brutto_MW.setText("{:.2f} €".format(total_cost_mw * 1.19))
        self.Reinigung_netto_RW.setText("{:.2f} €".format(total_cost_rw))
        self.Reinigung_brutto_RW.setText("{:.2f} €".format(total_cost_rw * 1.19))
        self.Reinigung_netto_SW.setText("{:.2f} €".format(total_cost_sw))
        self.Reinigung_brutto_SW.setText("{:.2f} €".format(total_cost_sw * 1.19))

    def calculateCleaningTVCost(self):
        self.load_cleaning_costs()
        dimension_column_index = 3
        length_column_index = 1
        system_column_index = 2
        GAL_column_index = 4
        cleaningTVcost_column_index = 6
        TVcost_column_index = 7
        GALcost_column_index = 8
        Panoramocost_column_index = 9

        total_cost_cleaning_tv_mw = 0
        total_cost_cleaning_tv_rw = 0
        total_cost_cleaning_tv_sw = 0

        system_mapping = {
            "Mischwasser": ["Mischwasser", "KM Mischwasserkanal"],
            "Regenwasser": ["Regenwasser", "KR Regenwasserkanal", "BA Bachverrohrung"],
            "Schmutzwasser": ["Schmutzwasser", "KS Schmutzwasserkanal"]
        }

        for row in range(self.Liste_Haltungen.rowCount()):
            dimension_item = self.Liste_Haltungen.item(row, dimension_column_index)
            length_item = self.Liste_Haltungen.item(row, length_column_index)
            system_item = self.Liste_Haltungen.item(row, system_column_index)

            if dimension_item is not None and length_item is not None and system_item is not None:
                dimension = dimension_item.text()
                length_text = length_item.text()
                system = system_item.text()

                if length_text is None or length_text.strip() == '' or length_text.lower() == 'none':
                    continue

                try:
                    length = float(length_text)
                except ValueError:
                    continue

                for key, values in system_mapping.items():
                    if system in values:
                        system = key
                        break

                if dimension in self.reinigung_TV_costs and system in self.reinigung_TV_costs[dimension]:
                    cost_per_unit = self.reinigung_TV_costs[dimension][system]
                    cost = length * cost_per_unit

                    brutto_cost = cost * 1.19
                    cost_item = QTableWidgetItem("{:.2f} €".format(brutto_cost))
                    self.Liste_Haltungen.setItem(row, cleaningTVcost_column_index, cost_item)

                    if system == "Mischwasser":
                        total_cost_cleaning_tv_mw += cost
                    elif system == "Regenwasser":
                        total_cost_cleaning_tv_rw += cost
                    elif system == "Schmutzwasser":
                        total_cost_cleaning_tv_sw += cost

        self.Reinigung_befahrung_netto_MW.setText("{:.2f} €".format(total_cost_cleaning_tv_mw))
        self.Reinigung_befahrung_brutto_MW.setText("{:.2f} €".format(total_cost_cleaning_tv_mw * 1.19))
        self.Reinigung_befahrung_netto_RW.setText("{:.2f} €".format(total_cost_cleaning_tv_rw))
        self.Reinigung_befahrung_brutto_RW.setText("{:.2f} €".format(total_cost_cleaning_tv_rw * 1.19))
        self.Reinigung_befahrung_netto_SW.setText("{:.2f} €".format(total_cost_cleaning_tv_sw))
        self.Reinigung_befahrung_brutto_SW.setText("{:.2f} €".format(total_cost_cleaning_tv_sw * 1.19))

        total_cost_tv_mw = 0
        total_cost_tv_rw = 0
        total_cost_tv_sw = 0
        total_cost_Panoramo_tv_mw = 0
        total_cost_Panoramo_tv_rw = 0
        total_cost_Panoramo_tv_sw = 0
        total_cost_GAL_mw = 0
        total_cost_GAL_rw = 0
        total_cost_GAL_sw = 0
        cost_Panoramo_SI = 37.51
        count_mw = 0
        count_rw = 0
        count_sw = 0

        for row in range(self.Liste_Haltungen.rowCount()):
            dimension_item = self.Liste_Haltungen.item(row, dimension_column_index)
            length_item = self.Liste_Haltungen.item(row, length_column_index)
            system_item = self.Liste_Haltungen.item(row, system_column_index)
            GAL_item = self.Liste_Haltungen.item(row, GAL_column_index)
            cost_Panoramo = 0

            if dimension_item is not None and length_item is not None and system_item is not None:
                dimension = dimension_item.text()
                length_text = length_item.text()
                system = system_item.text()

                if length_text is None or length_text.strip() == '' or length_text.lower() == 'none':
                    continue

                try:
                    length = float(length_text)
                except ValueError:
                    continue

                for key, values in system_mapping.items():
                    if system in values:
                        system = key
                        break

                GAL_value_text = GAL_item.text() if GAL_item else '0'
                try:
                    GAL_value_float = float(GAL_value_text) if GAL_value_text.strip().lower() != 'none' else 0.0
                except ValueError:
                    GAL_value_float = 0.0
                    
                if dimension in self.TV_costs and system in self.TV_costs[dimension]:
                    cost_per_unit = self.TV_costs[dimension][system]
                    cost = length * cost_per_unit
                    brutto_cost = cost * 1.19
                    cost_item = QTableWidgetItem("{:.2f} €".format(brutto_cost))
                    self.Liste_Haltungen.setItem(row, TVcost_column_index, cost_item)

                    if system == "Mischwasser":
                        total_cost_tv_mw += cost
                        count_mw += 1
                    elif system == "Regenwasser":
                        total_cost_tv_rw += cost
                        count_rw += 1
                    elif system == "Schmutzwasser":
                        total_cost_tv_sw += cost
                        count_sw += 1

                if self.checkBox_Panoramo.isChecked():
                    if dimension in self.Panoramo_costs and system in self.Panoramo_costs[dimension]:
                        cost_Panoramo_per_unit = self.Panoramo_costs[dimension][system]
                        cost_Panoramo = length * cost_Panoramo_per_unit
                        brutto_cost_Panoramo = (cost_Panoramo + cost_Panoramo_SI + cost) * 1.19
                        brutto_cost_Panoramo_only = (cost_Panoramo + cost_Panoramo_SI) * 1.19
                        cost_item_Panoramo_only = QTableWidgetItem("{:.2f} €".format(brutto_cost_Panoramo_only))
                        self.Liste_Haltungen.setItem(row, Panoramocost_column_index, cost_item_Panoramo_only)

                        if system == "Mischwasser":
                            total_cost_Panoramo_tv_mw += cost_Panoramo
                        elif system == "Regenwasser":
                            total_cost_Panoramo_tv_rw += cost_Panoramo
                        elif system == "Schmutzwasser":
                            total_cost_Panoramo_tv_sw += cost_Panoramo

                if self.checkBox_GAL.isChecked():
                    if dimension in self.TV_SAT_costs and system in self.TV_SAT_costs[dimension]:
                        cost_GAL_per_unit = self.TV_SAT_costs[dimension][system]
                        cost_GAL = GAL_value_float * cost_GAL_per_unit + GAL_value_float * 19.59
                        brutto_cost_GAL = (cost + cost_GAL) * 1.19
                        brutto_cost_GAL_only = (cost_GAL) * 1.19
                        cost_item_GAL_only = QTableWidgetItem("{:.2f} €".format(brutto_cost_GAL_only))
                        self.Liste_Haltungen.setItem(row, GALcost_column_index, cost_item_GAL_only)

                        if system == "Mischwasser":
                            total_cost_GAL_mw += cost_GAL
                        elif system == "Regenwasser":
                            total_cost_GAL_rw += cost_GAL
                        elif system == "Schmutzwasser":
                            total_cost_GAL_sw += cost_GAL

        total_cost_tv_pano_gal_mw = 0
        total_cost_tv_pano_gal_rw = 0
        total_cost_tv_pano_gal_sw = 0
        if self.checkBox_Panoramo.isChecked() and self.checkBox_GAL.isChecked():
            total_cost_tv_pano_gal_mw = total_cost_tv_mw + total_cost_Panoramo_tv_mw + (count_mw * cost_Panoramo_SI) + total_cost_GAL_mw
            total_cost_tv_pano_gal_rw = total_cost_tv_rw + total_cost_Panoramo_tv_rw + (count_rw * cost_Panoramo_SI) + total_cost_GAL_rw
            total_cost_tv_pano_gal_sw = total_cost_tv_sw + total_cost_Panoramo_tv_sw + (count_sw * cost_Panoramo_SI) + total_cost_GAL_sw

            self.Befahrung_netto_MW.setText("{:.2f} €".format(total_cost_tv_pano_gal_mw))
            self.Befahrung_brutto_MW.setText("{:.2f} €".format(total_cost_tv_pano_gal_mw * 1.19))
            self.Befahrung_netto_RW.setText("{:.2f} €".format(total_cost_tv_pano_gal_rw))
            self.Befahrung_brutto_RW.setText("{:.2f} €".format(total_cost_tv_pano_gal_rw * 1.19))
            self.Befahrung_netto_SW.setText("{:.2f} €".format(total_cost_tv_pano_gal_sw))
            self.Befahrung_brutto_SW.setText("{:.2f} €".format(total_cost_tv_pano_gal_sw * 1.19))

        elif self.checkBox_Panoramo.isChecked():
            total_cost_tv_mw += total_cost_Panoramo_tv_mw + (count_mw * cost_Panoramo_SI)
            total_cost_tv_rw += total_cost_Panoramo_tv_rw + (count_rw * cost_Panoramo_SI)
            total_cost_tv_sw += total_cost_Panoramo_tv_sw + (count_sw * cost_Panoramo_SI)

            self.Befahrung_netto_MW.setText("{:.2f} €".format(total_cost_tv_mw))
            self.Befahrung_brutto_MW.setText("{:.2f} €".format(total_cost_tv_mw * 1.19))
            self.Befahrung_netto_RW.setText("{:.2f} €".format(total_cost_tv_rw))
            self.Befahrung_brutto_RW.setText("{:.2f} €".format(total_cost_tv_rw * 1.19))
            self.Befahrung_netto_SW.setText("{:.2f} €".format(total_cost_tv_sw))
            self.Befahrung_brutto_SW.setText("{:.2f} €".format(total_cost_tv_sw * 1.19))

        elif self.checkBox_GAL.isChecked():
            total_cost_tv_mw += total_cost_GAL_mw
            total_cost_tv_rw += total_cost_GAL_rw
            total_cost_tv_sw += total_cost_GAL_sw

            self.Befahrung_netto_MW.setText("{:.2f} €".format(total_cost_tv_mw))
            self.Befahrung_brutto_MW.setText("{:.2f} €".format(total_cost_tv_mw * 1.19))
            self.Befahrung_netto_RW.setText("{:.2f} €".format(total_cost_tv_rw))
            self.Befahrung_brutto_RW.setText("{:.2f} €".format(total_cost_tv_rw * 1.19))
            self.Befahrung_netto_SW.setText("{:.2f} €".format(total_cost_tv_sw))
            self.Befahrung_brutto_SW.setText("{:.2f} €".format(total_cost_tv_sw * 1.19))

        else:
            self.Befahrung_netto_MW.setText("{:.2f} €".format(total_cost_tv_mw))
            self.Befahrung_brutto_MW.setText("{:.2f} €".format(total_cost_tv_mw * 1.19))
            self.Befahrung_netto_RW.setText("{:.2f} €".format(total_cost_tv_rw))
            self.Befahrung_brutto_RW.setText("{:.2f} €".format(total_cost_tv_rw * 1.19))
            self.Befahrung_netto_SW.setText("{:.2f} €".format(total_cost_tv_sw))
            self.Befahrung_brutto_SW.setText("{:.2f} €".format(total_cost_tv_sw * 1.19))

    def export_excel(self):
        data_frame = pd.DataFrame(columns=["Haltungsname", "Laenge", "Dimension", "Entwaesserungssystem", "Anzahl GALs", "Kosten Reinigung", "Kosten Reinigung TV", "Kosten TV", "Kosten GAL", "Kosten Panoramo"])

        for row in range(self.Liste_Haltungen.rowCount()):
            data_row = []
            for column in range(self.Liste_Haltungen.columnCount()):
                item = self.Liste_Haltungen.item(row, column)
                if item is not None:
                    if column == 1:
                        text = item.text().replace('.', ',')
                        data_row.append(text)
                    else:
                        data_row.append(item.text())
                else:
                    data_row.append('')

            while len(data_row) < len(data_frame.columns):
                data_row.append('')
            while len(data_row) > len(data_frame.columns):
                data_row.pop()

            data_frame.loc[len(data_frame)] = data_row

        template_options = QFileDialog.Options()
        template_file_path, _ = QFileDialog.getOpenFileName(None, "Excel-Vorlage auswählen", "", "Excel-Dateien (*.xlsm)", options=template_options)

        if not template_file_path: return

        workbook = openpyxl.load_workbook(template_file_path, keep_vba=True)
        sheet = workbook.active

        for row in dataframe_to_rows(data_frame, index=False, header=True):
            sheet.append(row)

        for row_idx, row_data in enumerate(data_frame.values, 2):
            for col_idx, cell_data in enumerate(row_data):
                cell = sheet.cell(row=row_idx, column=col_idx + 1)
                if isinstance(cell_data, (int, float)):
                    cell.number_format = '#,##0.00 [$€-407];[Red]-#,##0.00 [$€-407]'

        export_options = QFileDialog.Options()
        export_file_path, _ = QFileDialog.getSaveFileName(None, "Excel-Datei speichern", "", "Excel-Dateien (*.xlsm)", options=export_options)

        if export_file_path:
            workbook.save(export_file_path)
            try:
                if os.name == 'nt':
                    os.startfile(export_file_path)
                else:
                    subprocess.Popen(['xdg-open', export_file_path])
            except Exception as e:
                print(f"Fehler beim Öffnen der Datei: {e}")

